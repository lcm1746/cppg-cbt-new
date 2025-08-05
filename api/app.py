import os
import random
import json
from openpyxl import load_workbook

def extract_questions_from_excel(file_path):
    """openpyxl을 사용해서 엑셀 파일에서 문제 추출"""
    try:
        # openpyxl로 엑셀 파일 읽기
        workbook = load_workbook(filename=file_path, read_only=True)
        worksheet = workbook.active
        
        print(f"엑셀 파일 읽기 완료")
        
        questions_by_set = {}
        
        # 헤더 행 찾기
        headers = []
        for row in worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(cell) if cell else '' for cell in row]
            break
        
        print(f"컬럼: {headers}")
        
        # 데이터 행 처리
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            question = {}
            
            # 세트 처리
            set_idx = headers.index('세트') if '세트' in headers else -1
            if set_idx >= 0 and row[set_idx]:
                set_num = int(row[set_idx])
            else:
                set_num = 1
                
            # 문제 번호 처리
            num_idx = headers.index('문제번호') if '문제번호' in headers else -1
            if num_idx >= 0 and row[num_idx]:
                question['number'] = int(row[num_idx])
            else:
                question['number'] = row_idx - 1
                
            # 문제 내용 처리
            text_idx = headers.index('문제') if '문제' in headers else 0
            if text_idx >= 0 and row[text_idx]:
                question['text'] = str(row[text_idx])
            else:
                question['text'] = str(row[0]) if row[0] else ''
                
            # 보기 처리
            options_idx = headers.index('보기') if '보기' in headers else -1
            if options_idx >= 0 and row[options_idx]:
                options_text = str(row[options_idx])
                options = []
                for i in range(1, 6):
                    option_pattern = f'{chr(9311 + i)}'
                    if option_pattern in options_text:
                        start = options_text.find(option_pattern)
                        if i < 5:
                            end = options_text.find(f'{chr(9311 + i + 1)}')
                            if end == -1:
                                end = len(options_text)
                        else:
                            end = len(options_text)
                        option_text = options_text[start+1:end].strip()
                        options.append(option_text)
                    else:
                        options.append(f"보기 {i}")
                
                while len(options) < 5:
                    options.append(f"보기 {len(options) + 1}")
            else:
                options = ["보기 1", "보기 2", "보기 3", "보기 4", "보기 5"]
            
            question['options'] = options[:5]
            
            # 정답 처리
            answer_idx = headers.index('답안') if '답안' in headers else -1
            if answer_idx >= 0 and row[answer_idx]:
                answer = str(row[answer_idx])
                answer_map = {'①': 0, '②': 1, '③': 2, '④': 3, '⑤': 4, '1': 0, '2': 1, '3': 2, '4': 3, '5': 4}
                question['correct'] = answer_map.get(answer, 0)
            else:
                question['correct'] = 0
                
            # 해설 처리
            explanation_idx = headers.index('해설') if '해설' in headers else -1
            if explanation_idx >= 0 and row[explanation_idx]:
                question['answer'] = str(row[explanation_idx])
            else:
                question['answer'] = f"문제 {question['number']}의 정답은 {question['options'][question['correct']]}입니다."
            
            # 세트별로 문제 분류
            if set_num not in questions_by_set:
                questions_by_set[set_num] = []
            
            # 빈 문제 제외
            if question['text'].strip():
                questions_by_set[set_num].append(question)
        
        workbook.close()
        return questions_by_set
        
    except Exception as e:
        print(f"엑셀 파일 처리 오류: {e}")
        return None

def load_questions_from_file():
    """CPPG 파일에서 문제 로드"""
    file_paths = [
        "cppg_qa_final.xlsx",
        "../cppg_qa_final.xlsx",
        "./cppg_qa_final.xlsx"
    ]
    
    file_path = None
    for path in file_paths:
        if os.path.exists(path):
            file_path = path
            break
    
    if not file_path:
        print(f"파일을 찾을 수 없습니다. 시도한 경로: {file_paths}")
        return None, None
    
    print(f"파일 로드 중: {file_path}")
    questions_by_set = extract_questions_from_excel(file_path)
    
    if questions_by_set:
        stats = {}
        total_questions = 0
        for set_num, questions in questions_by_set.items():
            stats[f'세트{set_num}'] = len(questions)
            total_questions += len(questions)
        stats['총문제수'] = total_questions
        
        return questions_by_set, stats
    
    return None, None

# 전역 변수로 문제 로드
QUESTIONS_BY_SET = None
STATS = None

def handler(request, context):
    """Vercel 서버리스 함수"""
    try:
        # 요청 정보 추출
        path = request.get('path', '/')
        method = request.get('method', 'GET')
        
        # 문제 로드 (한 번만)
        global QUESTIONS_BY_SET, STATS
        if QUESTIONS_BY_SET is None:
            QUESTIONS_BY_SET, STATS = load_questions_from_file()
        
        if path == '/' and method == 'GET':
            if QUESTIONS_BY_SET is None:
                return {
                    'statusCode': 200,
                    'headers': {'Content-Type': 'text/html; charset=utf-8'},
                    'body': '<h1>문제 파일을 로드할 수 없습니다.</h1>'
                }
            
            # 메인 페이지 HTML
            html = f'''
            <!DOCTYPE html>
            <html>
            <head>
                <title>CPPG CBT 시스템</title>
                <meta charset="utf-8">
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 40px; background: #f5f5f5; }}
                    .container {{ max-width: 800px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                    .btn {{ display: inline-block; padding: 15px 30px; margin: 10px; 
                           background: #007bff; color: white; text-decoration: none; 
                           border-radius: 5px; font-weight: bold; transition: background 0.3s; }}
                    .btn:hover {{ background: #0056b3; }}
                    .stats {{ background: #f8f9fa; padding: 20px; border-radius: 5px; margin: 20px 0; border-left: 4px solid #007bff; }}
                    h1 {{ color: #333; text-align: center; }}
                    .stats h3 {{ color: #007bff; margin-top: 0; }}
                </style>
            </head>
            <body>
                <div class="container">
                    <h1>CPPG CBT 시스템</h1>
                    <div class="stats">
                        <h3>📊 문제 통계</h3>
                        <p><strong>총 문제수:</strong> {STATS.get("총문제수", 0)}개</p>
                        <p><strong>세트1:</strong> {STATS.get("세트1", 0)}개</p>
                        <p><strong>세트2:</strong> {STATS.get("세트2", 0)}개</p>
                        <p><strong>세트3:</strong> {STATS.get("세트3", 0)}개</p>
                        <p><strong>세트4:</strong> {STATS.get("세트4", 0)}개</p>
                        <p><strong>세트5:</strong> {STATS.get("세트5", 0)}개</p>
                    </div>
                    <div style="text-align: center;">
                        <a href="/practice" class="btn">📚 순차 연습</a>
                        <a href="/random" class="btn">🎲 랜덤 연습</a>
                        <a href="/exam" class="btn">📝 시험 모드</a>
                    </div>
                </div>
            </body>
            </html>
            '''
            
            return {
                'statusCode': 200,
                'headers': {'Content-Type': 'text/html; charset=utf-8'},
                'body': html
            }
        
        elif path == '/api/questions' and method == 'GET':
            if not QUESTIONS_BY_SET:
                return {
                    'statusCode': 400,
                    'headers': {'Content-Type': 'application/json'},
                    'body': json.dumps({'error': '문제가 로드되지 않았습니다.'})
                }
            
            all_questions = []
            for set_num, questions in QUESTIONS_BY_SET.items():
                for question in questions:
                    question_copy = question.copy()
                    question_copy['set'] = set_num
                    all_questions.append(question_copy)
            
            return {
                'statusCode': 200,
                'headers': {'Content-Type': 'application/json'},
                'body': json.dumps({'questions': all_questions})
            }
        
        elif path == '/api/exam-questions' and method == 'GET':
            if not QUESTIONS_BY_SET:
                return {
                    'statusCode': 400,
                    'headers': {'Content-Type': 'application/json'},
                    'body': json.dumps({'error': '문제가 로드되지 않았습니다.'})
                }
            
            # 시험 문제 생성
            exam_questions = []
            for set_num, questions in QUESTIONS_BY_SET.items():
                if questions:
                    selected = random.sample(questions, min(10, len(questions)))
                    for i, question in enumerate(selected):
                        question_copy = question.copy()
                        question_copy['exam_number'] = len(exam_questions) + 1
                        question_copy['set'] = set_num
                        exam_questions.append(question_copy)
            
            return {
                'statusCode': 200,
                'headers': {'Content-Type': 'application/json'},
                'body': json.dumps({'questions': exam_questions})
            }
        
        else:
            return {
                'statusCode': 404,
                'headers': {'Content-Type': 'text/html; charset=utf-8'},
                'body': '<h1>404 - 페이지를 찾을 수 없습니다</h1>'
            }
            
    except Exception as e:
        print(f"Error in handler: {e}")
        return {
            'statusCode': 500,
            'headers': {'Content-Type': 'text/html; charset=utf-8'},
            'body': f'<h1>Internal Server Error: {str(e)}</h1>'
        } 